from requests import get
from requests.exceptions import RequestException
from contextlib import closing
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os.path
import json

# Sends a get request to the given url, if response returns 200 status code returns the content, otherwise returns null

def simple_get(url):
    try:
        with closing(get(url, stream=True)) as resp:
            if is_good_response(resp):
                return resp.content
            else:
                return None

    except RequestException as e:
        log_error('Error during requests to {0} : {1}'.format(url, str(e)))
        return None

# Returns true if response returns 200 status code and contains HTML

def is_good_response(resp):
    content_type = resp.headers['Content-Type'].lower()
    return (resp.status_code == 200
            and content_type is not None
            and content_type.find('html') > -1)

# Prints the given error. It can be modified to behave differently on specific errors

def log_error(e):
    print(e)

# Returns the closing price for the specified symbol

def getPrice(symbol):
    raw_html = simple_get(f'https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={symbol}')
    html = BeautifulSoup(raw_html, 'html.parser')
    ul = html.select('#MainContent_PanelInfo > div.main-indicators > ul.top-list > li:nth-child(1) > span')
    price = str(ul[0]).replace("<span>", "")
    price = price.replace("</span>", "")
    price = price.replace(",", ".")
    return float(price)

# Fills the headers of the table in Excel worksheet

def fillColumns(worksheet):
    worksheet["A1"] = "Symbol"
    worksheet["B1"] = "Current Price"
    worksheet["C1"] = "Previous Price"
    worksheet["D1"] = "Change(%)"

# Load the symbols from symbols.json file then creates/opens Portfolio.xlsx and set it as active worksheet

with open('symbols.json') as file:
    symbols = json.load(file)['symbols']
if os.path.exists('./Portfolio.xlsx'):
    wb = load_workbook('./Portfolio.xlsx')
else:
    wb = Workbook()
ws = wb.active
fillColumns(ws)

# Write symbol name, current closing price, previous closing price, change in % to worksheet
# TODO: Add conditional formatting to color the cells on increase/decrease

for index in range(len(symbols)):
    ws["A" + str(1 + index + 1)] = symbols[index]
    if ws["B" + str(1 + index + 1)] != "":
        ws["C" + str(1 + index + 1)] = ws["B" + str(1 + index + 1)].value
        ws["B" + str(1 + index + 1)] = getPrice(symbols[index])
        ws["D" + str(1 + index + 1)] = f'=(B{str(1 + index + 1)}-C{str(1 + index + 1)})/C{str(1 + index + 1)}*100'
    else:
        ws["B" + str(1 + index + 1)] = getPrice(symbols[index]) # Insert prices to B column
wb.save('./Portfolio.xlsx')
